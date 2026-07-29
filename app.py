import streamlit as st
import json
import os
import hashlib
import io
import base64
import pandas as pd
from datetime import date
from PIL import Image, ImageOps

# Configuración principal de la app
st.set_page_config(page_title="Sistema de Inspección de Tolvas CAT", layout="wide")

st.title("📋 Reporte de Inspección de Tolvas CAT 794 AC")

# --- COMPONENTE DE ANOTACIÓN DE FOTOS ---
_ANOTADOR_HTML = """
<div>
  <div id="barra" style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:6px;">
    <button data-tool="none" class="sel">🚫 Ninguno</button>
    <button data-tool="circle">⭕ Círculo</button>
    <button data-tool="rect">🔲 Rectángulo</button>
    <button data-tool="line">↗️ Flecha</button>
    <button data-tool="freedraw">✏️ Libre</button>
    <button id="deshacer">↩️ Deshacer</button>
    <button id="limpiar">🗑️ Borrar todo</button>
    <button id="guardar" class="guardar">💾 Guardar anotación</button>
  </div>
  <canvas id="lienzo" width="600" height="450"></canvas>
  <div id="estado"></div>
</div>
"""

_ANOTADOR_CSS = """
#barra button {
  background:#262730; color:#fafafa; border:1px solid #555; border-radius:6px;
  padding:7px 10px; font-size:13px; cursor:pointer;
}
#barra button:hover { background:#3a3b45; }
#barra button.sel { background:#FF4B4B; border-color:#FF4B4B; }
#barra button.guardar { background:#21c45d; border-color:#21c45d; font-weight:bold; }
canvas { border:1px solid #444; width:100%; height:auto; touch-action:pan-y; display:block; }
#estado { font-size:12px; color:#9a9a9a; margin-top:4px; }
"""

_ANOTADOR_JS = """
export default function(component) {
  const { data, parentElement, setStateValue } = component;

  if (!parentElement._anotadorInit) {
    parentElement._anotadorInit = true;
    parentElement._formas = [];
    parentElement._herramienta = "none";
    parentElement._imagenFondo = null;

    const canvas = parentElement.querySelector('#lienzo');
    const ctx = canvas.getContext('2d');
    const estadoEl = parentElement.querySelector('#estado');

    function redibujar() {
      ctx.clearRect(0, 0, canvas.width, canvas.height);
      if (parentElement._imagenFondo) ctx.drawImage(parentElement._imagenFondo, 0, 0, canvas.width, canvas.height);
      const todas = parentElement._formas.concat(parentElement._formaActual ? [parentElement._formaActual] : []);
      todas.forEach(dibujarForma);
    }
    function dibujarForma(f) {
      if (!f) return;
      ctx.lineWidth = 3; ctx.strokeStyle = "#FF0000"; ctx.fillStyle = "rgba(255,0,0,0.2)";
      if (f.tipo === "circle") {
        const cx=(f.x1+f.x2)/2, cy=(f.y1+f.y2)/2, rx=Math.abs(f.x2-f.x1)/2, ry=Math.abs(f.y2-f.y1)/2;
        ctx.beginPath(); ctx.ellipse(cx,cy,rx,ry,0,0,2*Math.PI); ctx.fill(); ctx.stroke();
      } else if (f.tipo === "rect") {
        ctx.beginPath(); ctx.rect(f.x1,f.y1,f.x2-f.x1,f.y2-f.y1); ctx.fill(); ctx.stroke();
      } else if (f.tipo === "line") {
        const ang = Math.atan2(f.y2-f.y1, f.x2-f.x1);
        ctx.beginPath(); ctx.moveTo(f.x1,f.y1); ctx.lineTo(f.x2,f.y2); ctx.stroke();
        const t=12;
        ctx.beginPath(); ctx.moveTo(f.x2,f.y2);
        ctx.lineTo(f.x2-t*Math.cos(ang-Math.PI/6), f.y2-t*Math.sin(ang-Math.PI/6));
        ctx.lineTo(f.x2-t*Math.cos(ang+Math.PI/6), f.y2-t*Math.sin(ang+Math.PI/6));
        ctx.closePath(); ctx.fillStyle="#FF0000"; ctx.fill();
      } else if (f.tipo === "freedraw") {
        ctx.beginPath();
        f.puntos.forEach((p,i)=> i===0 ? ctx.moveTo(p.x,p.y) : ctx.lineTo(p.x,p.y));
        ctx.stroke();
      }
    }
    function coords(ev) {
      const r = canvas.getBoundingClientRect();
      const ex = ev.touches ? ev.touches[0].clientX : ev.clientX;
      const ey = ev.touches ? ev.touches[0].clientY : ev.clientY;
      return { x: (ex-r.left)*(canvas.width/r.width), y: (ey-r.top)*(canvas.height/r.height) };
    }
    function iniciar(ev) {
      if (parentElement._herramienta === "none") return;
      ev.preventDefault();
      parentElement._dibujando = true;
      const {x,y} = coords(ev);
      parentElement._formaActual = parentElement._herramienta === "freedraw"
        ? {tipo:"freedraw",puntos:[{x,y}]}
        : {tipo:parentElement._herramienta,x1:x,y1:y,x2:x,y2:y};
    }
    function mover(ev) {
      if (!parentElement._dibujando) return;
      ev.preventDefault();
      const {x,y} = coords(ev);
      if (parentElement._herramienta === "freedraw") parentElement._formaActual.puntos.push({x,y});
      else { parentElement._formaActual.x2 = x; parentElement._formaActual.y2 = y; }
      redibujar();
    }
    function terminar() {
      if (!parentElement._dibujando) return;
      parentElement._dibujando = false;
      if (parentElement._formaActual) parentElement._formas.push(parentElement._formaActual);
      parentElement._formaActual = null;
      redibujar();
      estadoEl.textContent = "Sin guardar todavía";
    }

    canvas.addEventListener("mousedown", iniciar);
    canvas.addEventListener("mousemove", mover);
    window.addEventListener("mouseup", terminar);
    canvas.addEventListener("touchstart", iniciar, {passive:false});
    canvas.addEventListener("touchmove", mover, {passive:false});
    canvas.addEventListener("touchend", terminar);

    parentElement.querySelectorAll('#barra button[data-tool]').forEach((btn) => {
      btn.addEventListener("click", () => {
        parentElement.querySelectorAll('#barra button[data-tool]').forEach((b) => b.classList.remove("sel"));
        btn.classList.add("sel");
        parentElement._herramienta = btn.dataset.tool;
        canvas.style.touchAction = (btn.dataset.tool === "none") ? "pan-y" : "none";
      });
    });
    parentElement.querySelector('#deshacer').addEventListener("click", () => { parentElement._formas.pop(); redibujar(); });
    parentElement.querySelector('#limpiar').addEventListener("click", () => { parentElement._formas = []; redibujar(); });
    parentElement.querySelector('#guardar').addEventListener("click", () => {
      setStateValue('imagen_anotada', canvas.toDataURL('image/png'));
      estadoEl.textContent = "✅ Anotación guardada";
    });

    parentElement._redibujar = redibujar;
  }

  if (data && data.imagen_base64 && parentElement._ultimaImagen !== data.imagen_base64) {
    parentElement._ultimaImagen = data.imagen_base64;
    const canvasEl = parentElement.querySelector('#lienzo');
    const img = new Image();
    img.onload = () => {
      canvasEl.width = img.naturalWidth;
      canvasEl.height = img.naturalHeight;
      parentElement._imagenFondo = img;
      parentElement._redibujar();
    };
    img.src = "data:image/png;base64," + data.imagen_base64;
  }
}
"""

_componente_anotador = st.components.v2.component(
    "anotador_fotos_tolva",
    html=_ANOTADOR_HTML,
    css=_ANOTADOR_CSS,
    js=_ANOTADOR_JS,
)

def anotador_fotos(imagen_base64_sin_prefijo, key):
    resultado = _componente_anotador(
        data={"imagen_base64": imagen_base64_sin_prefijo},
        key=key,
        on_imagen_anotada_change=lambda: None
    )
    return resultado.imagen_anotada if resultado else None

# --- COMPONENTE DE CÁMARA CON ZOOM NATIVO ---
_CAMARA_HTML = """
<div>
  <video id="video" autoplay playsinline muted></video>
  <div id="controlesZoom" style="display:none;">
    <span>🔍</span>
    <input type="range" id="zoomSlider" min="1" max="8" step="0.1" value="1">
    <span id="zoomValor">1.0x</span>
  </div>
  <div style="display:flex;gap:8px;margin-top:8px;">
    <button id="btnCapturar" class="capturar">📸 Capturar Foto</button>
  </div>
  <div id="estadoCam" style="font-size:12px;color:#9a9a9a;margin-top:4px;"></div>
</div>
"""

_CAMARA_CSS = """
video { width:100%; border-radius:8px; background:#000; touch-action:none; display:block; }
#controlesZoom { display:flex; align-items:center; gap:8px; margin-top:6px; color:#fafafa; font-size:13px; }
#controlesZoom input[type=range] { flex:1; }
button.capturar {
  background:#21c45d; color:#fff; border:none; border-radius:6px;
  padding:10px 16px; font-size:14px; font-weight:bold; cursor:pointer;
}
"""

_CAMARA_JS = """
export default function(component) {
  const { parentElement, setStateValue } = component;
  if (parentElement._camInit) return;
  parentElement._camInit = true;

  const video = parentElement.querySelector('#video');
  const estadoEl = parentElement.querySelector('#estadoCam');
  const controlesZoom = parentElement.querySelector('#controlesZoom');
  const zoomSlider = parentElement.querySelector('#zoomSlider');
  const zoomValor = parentElement.querySelector('#zoomValor');
  let track = null;
  let usaZoomHardware = false;
  let zoomDigital = 1;

  navigator.mediaDevices.getUserMedia({
    video: { facingMode: { ideal: "environment" }, width: { ideal: 1920 }, height: { ideal: 1080 } },
    audio: false
  }).then((stream) => {
    video.srcObject = stream;
    track = stream.getVideoTracks()[0];
    const capacidades = track.getCapabilities ? track.getCapabilities() : {};
    if (capacidades.zoom) {
      usaZoomHardware = true;
      zoomSlider.min = capacidades.zoom.min;
      zoomSlider.max = capacidades.zoom.max;
      zoomSlider.step = capacidades.zoom.step || 0.1;
      zoomSlider.value = capacidades.zoom.min;
      controlesZoom.style.display = "flex";
    } else {
      controlesZoom.style.display = "flex";
    }
    estadoEl.textContent = "Cámara lista";
  }).catch((err) => {
    estadoEl.textContent = "⚠️ No se pudo abrir la cámara: " + err.message;
  });

  function aplicarZoom(valor) {
    if (usaZoomHardware && track) {
      track.applyConstraints({ advanced: [{ zoom: parseFloat(valor) }] }).catch(() => {});
    } else {
      zoomDigital = parseFloat(valor);
      video.style.transform = "scale(" + zoomDigital + ")";
    }
    zoomValor.textContent = parseFloat(valor).toFixed(1) + "x";
  }
  zoomSlider.addEventListener("input", (ev) => aplicarZoom(ev.target.value));

  let distanciaInicial = null, zoomInicial = 1;
  video.addEventListener("touchstart", (ev) => {
    if (ev.touches.length === 2) {
      distanciaInicial = Math.hypot(
        ev.touches[0].clientX - ev.touches[1].clientX,
        ev.touches[0].clientY - ev.touches[1].clientY
      );
      zoomInicial = parseFloat(zoomSlider.value);
    }
  }, { passive: true });
  video.addEventListener("touchmove", (ev) => {
    if (ev.touches.length === 2 && distanciaInicial) {
      const distanciaActual = Math.hypot(
        ev.touches[0].clientX - ev.touches[1].clientX,
        ev.touches[0].clientY - ev.touches[1].clientY
      );
      const factor = distanciaActual / distanciaInicial;
      let nuevoZoom = zoomInicial * factor;
      const min = parseFloat(zoomSlider.min), max = parseFloat(zoomSlider.max);
      nuevoZoom = Math.max(min, Math.min(max, nuevoZoom));
      zoomSlider.value = nuevoZoom;
      aplicarZoom(nuevoZoom);
    }
  }, { passive: true });

  parentElement.querySelector('#btnCapturar').addEventListener("click", () => {
    const lienzoCaptura = document.createElement("canvas");
    if (usaZoomHardware || zoomDigital === 1) {
      lienzoCaptura.width = video.videoWidth;
      lienzoCaptura.height = video.videoHeight;
      lienzoCaptura.getContext("2d").drawImage(video, 0, 0);
    } else {
      const anchoRecorte = video.videoWidth / zoomDigital;
      const altoRecorte = video.videoHeight / zoomDigital;
      const x = (video.videoWidth - anchoRecorte) / 2;
      const y = (video.videoHeight - altoRecorte) / 2;
      lienzoCaptura.width = anchoRecorte;
      lienzoCaptura.height = altoRecorte;
      lienzoCaptura.getContext("2d").drawImage(video, x, y, anchoRecorte, altoRecorte, 0, 0, anchoRecorte, altoRecorte);
    }
    const resultado = lienzoCaptura.toDataURL("image/jpeg", 0.92);
    setStateValue("foto_capturada", resultado);
    estadoEl.textContent = "✅ Foto capturada";
    if (track) track.stop();
  });
}
"""

_componente_camara = st.components.v2.component(
    "camara_nativa_tolva",
    html=_CAMARA_HTML,
    css=_CAMARA_CSS,
    js=_CAMARA_JS,
)

def camara_nativa(key):
    resultado = _componente_camara(key=key, on_foto_capturada_change=lambda: None)
    return resultado.foto_capturada if resultado else None


_FILAS_ENCABEZADO_ZONAS_EXCEL = [30, 70, 96, 131, 162, 196]
_FILAS_MARCADOR_FOTOS_EXCEL = [43, 77, 109, 140, 176, 204]

def _insertar_filas_seguro(ws, fila_insercion, cantidad):
    merges_originales = list(ws.merged_cells.ranges)
    for mc in merges_originales:
        ws.unmerge_cells(str(mc))
    ws.insert_rows(fila_insercion, amount=cantidad)
    for mc in merges_originales:
        min_row, min_col, max_row, max_col = mc.min_row, mc.min_col, mc.max_row, mc.max_col
        if min_row >= fila_insercion:
            min_row += cantidad
            max_row += cantidad
        elif max_row >= fila_insercion:
            max_row += cantidad
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

def _copiar_estilo_bloque(ws, fila_ini_origen, fila_fin_origen, fila_ini_destino, max_col=17):
    import copy
    alto = fila_fin_origen - fila_ini_origen + 1
    for offset in range(alto):
        fila_o = fila_ini_origen + offset
        fila_d = fila_ini_destino + offset
        dim_o = ws.row_dimensions.get(fila_o)
        if dim_o and dim_o.height:
            ws.row_dimensions[fila_d].height = dim_o.height
        for col in range(1, max_col + 1):
            c_o = ws.cell(row=fila_o, column=col)
            c_d = ws.cell(row=fila_d, column=col)
            c_d.font = copy.copy(c_o.font)
            c_d.border = copy.copy(c_o.border)
            c_d.fill = copy.copy(c_o.fill)
            c_d.alignment = copy.copy(c_o.alignment)
            c_d.number_format = c_o.number_format

def _obtener_bloques_fotos(ws, fila_marcador):
    bloques = []
    fila_actual = fila_marcador + 1
    while True:
        rango_encontrado = None
        for mc in ws.merged_cells.ranges:
            if mc.min_col == 1 and mc.min_row == fila_actual:
                rango_encontrado = (mc.min_row, mc.max_row)
                break
        if rango_encontrado is None:
            break
        bloques.append(rango_encontrado)
        fila_actual = rango_encontrado[1] + 1
    return bloques

def _ancho_columnas_px(ws, col_ini, col_fin):
    from openpyxl.utils import get_column_letter
    total = 0
    for col in range(col_ini, col_fin + 1):
        letra = get_column_letter(col)
        dim = ws.column_dimensions.get(letra)
        ancho_unidades = dim.width if (dim and dim.width) else 8.43
        total += ancho_unidades * 7 + 5
    return int(total)

def _alto_filas_px(ws, fila_ini, fila_fin):
    total = 0
    for fila in range(fila_ini, fila_fin + 1):
        dim = ws.row_dimensions.get(fila)
        alto_puntos = dim.height if (dim and dim.height) else 15
        total += alto_puntos * 1333 / 1000
    return int(total)

def _preparar_imagen_para_insertar(imagen_pil, ancho_max_px, alto_max_px):
    img_copia = imagen_pil.copy().convert("RGB")
    img_copia.thumbnail((max(ancho_max_px, 10), max(alto_max_px, 10)), Image.LANCZOS)
    buf = io.BytesIO()
    img_copia.save(buf, format="PNG")
    return buf.getvalue(), img_copia.width, img_copia.height

def _desplazar_filas_drawing_xml(xml_texto, fila_insercion_0idx, cantidad):
    import re
    def reemplazar(m):
        fila = int(m.group(2))
        if fila >= fila_insercion_0idx:
            fila += cantidad
        return f"{m.group(1)}{fila}{m.group(3)}"
    return re.sub(r'(<xdr:row>)(\d+)(</xdr:row>)', reemplazar, xml_texto)

def _construir_anchor_imagen_xml(id_imagen, col_0idx, fila_0idx, ancho_px, alto_px, rid):
    emu_x = int(ancho_px * 9525)
    emu_y = int(alto_px * 9525)
    return (
        f'<xdr:oneCellAnchor>'
        f'<xdr:from><xdr:col>{col_0idx}</xdr:col><xdr:colOff>9525</xdr:colOff>'
        f'<xdr:row>{fila_0idx}</xdr:row><xdr:rowOff>9525</xdr:rowOff></xdr:from>'
        f'<xdr:ext cx="{emu_x}" cy="{emu_y}"/>'
        f'<xdr:pic>'
        f'<xdr:nvPicPr><xdr:cNvPr id="{id_imagen}" name="FotoApp{id_imagen}"/>'
        f'<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr></xdr:nvPicPr>'
        f'<xdr:blipFill><a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:embed="{rid}"/>'
        f'<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
        f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{emu_x}" cy="{emu_y}"/></a:xfrm>'
        f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
        f'</xdr:pic><xdr:clientData/></xdr:oneCellAnchor>'
    )

def generar_reporte_excel(ruta_plantilla, cliente, lugar, fecha_insp, cod_equipo,
                           cod_tolva, horometro, cod_informe, revision, pm,
                           estructura_zonas, nombre_realizado, fecha_firma, firma_archivo,
                           todos_los_rechazos, matriz_espesores=None):
    
    import openpyxl
    import zipfile

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb["TOLVA DT"]

    ws["C5"] = cliente
    ws["C6"] = lugar
    ws["C7"] = fecha_insp
    ws["H6"] = cod_equipo
    ws["H7"] = cod_tolva
    ws["L6"] = horometro
    ws["P5"] = cod_informe
    ws["P6"] = revision
    ws["P7"] = pm

    if pm in ["1000H", "2000H"] and matriz_espesores is not None:
        fila_inicio_matriz = 15 
        col_inicio_matriz = 3   
        
        for i in range(8):
            for j in range(7):
                valor = matriz_espesores.iloc[i, j]
                ws.cell(row=fila_inicio_matriz + i, column=col_inicio_matriz + j, value=valor)
                
        if (matriz_espesores == "-").all().all():
            ws.cell(row=fila_inicio_matriz - 1, column=col_inicio_matriz, value="NO SE REALIZÓ MEDICIÓN DE ESPESORES")

    desplazamiento = 0
    puntos_insercion = []   
    fotos_pendientes = []   

    for idx_z, bloque_zona in enumerate(estructura_zonas):
        fila_header = _FILAS_ENCABEZADO_ZONAS_EXCEL[idx_z] + desplazamiento
        fila_inicio_items = fila_header + 2

        for idx_i, item in enumerate(bloque_zona["items"]):
            cod_z, desc_z, tec_def = item
            key_id = f"z{idx_z}_{idx_i}"
            fila = fila_inicio_items + idx_i

            defecto = st.session_state.get(f"def_{key_id}", "LF")
            es_lf = (defecto == "LF")

            if es_lf:
                longitud = "-"
                est_post = "-"
                condicion = "ACEPTABLE"
                comentario = "-"
            else:
                opc_long = st.session_state.get(f"opclong_{key_id}", "Manual")
                if opc_long == "VARIOS":
                    longitud = "VARIOS"
                else:
                    longitud = st.session_state.get(f"longval_{key_id}", "100")
                est_post = st.session_state.get(f"est_{key_id}", "NR")
                condicion = "RECHAZADO"
                comentario = st.session_state.get(f"com_{key_id}", "CREAR OT")

            tecnica = st.session_state.get(f"tec_{key_id}", tec_def)

            ws.cell(row=fila, column=5, value=fecha_insp)
            ws.cell(row=fila, column=7, value=defecto)
            ws.cell(row=fila, column=8, value=longitud)
            ws.cell(row=fila, column=9, value=est_post)
            ws.cell(row=fila, column=11, value=tecnica)
            ws.cell(row=fila, column=12, value=condicion)
            ws.cell(row=fila, column=14, value=comentario)

        rechazos_zona = [r for r in todos_los_rechazos if r["key_id"].startswith(f"z{idx_z}_")]
        
        if idx_z == 1:
            rechazos_zona.append({"zona": "2", "descripcion": "LETRERO LATERAL RH", "defecto": "-", "key_id": "esp_z2_rh"})
            rechazos_zona.append({"zona": "2", "descripcion": "LETRERO LATERAL LH", "defecto": "-", "key_id": "esp_z2_lh"})
        elif idx_z == 5:
            rechazos_zona.append({"zona": "8", "descripcion": "LETRERO POSTERIOR", "defecto": "-", "key_id": "esp_z8_post"})

        fila_marcador = _FILAS_MARCADOR_FOTOS_EXCEL[idx_z] + desplazamiento
        bloques = _obtener_bloques_fotos(ws, fila_marcador)

        while len(bloques) < len(rechazos_zona):
            fila_ini_ultimo, fila_fin_ultimo = bloques[-1]
            alto_bloque = fila_fin_ultimo - fila_ini_ultimo + 1
            fila_insercion = fila_fin_ultimo + 1
            _insertar_filas_seguro(ws, fila_insercion, alto_bloque)
            _copiar_estilo_bloque(ws, fila_ini_ultimo, fila_fin_ultimo, fila_insercion)
            puntos_insercion.append((fila_insercion, alto_bloque))
            desplazamiento += alto_bloque
            bloques.append((fila_insercion, fila_insercion + alto_bloque - 1))

        for idx_b, (fila_ini, fila_fin) in enumerate(bloques):
            if idx_b >= len(rechazos_zona):
                break
            rechazo = rechazos_zona[idx_b]
            key_id = rechazo["key_id"]

            texto_desc = f"ZONA {rechazo['zona']}\n{rechazo['descripcion'].upper()}\n\n{rechazo['defecto']}"
            ws.cell(row=fila_ini, column=1, value=texto_desc)

            ancho_pano_px = _ancho_columnas_px(ws, 5, 11) - 6
            ancho_det_px = _ancho_columnas_px(ws, 12, 17) - 6
            alto_bloque_px = _alto_filas_px(ws, fila_ini, fila_fin) - 6

            for prefijo_foto, col_0idx, ancho_disponible in (
                ("pano", 4, ancho_pano_px), ("det", 11, ancho_det_px)
            ):
                llave_base = f"img_{prefijo_foto}_{key_id}"
                foto_anotada = st.session_state.get(f"{llave_base}_anotada")
                foto_original = st.session_state.get(llave_base)
                
                img_foto = None
                if foto_anotada:
                    datos_bin = base64.b64decode(foto_anotada.split(",", 1)[1])
                    img_foto = Image.open(io.BytesIO(datos_bin))
                elif foto_original is not None:
                    img_foto = foto_original
                    
                if img_foto is not None:
                    bytes_png, ancho_f, alto_f = _preparar_imagen_para_insertar(
                        img_foto, ancho_disponible, alto_bloque_px
                    )
                    fotos_pendientes.append({
                        "fila_0idx": fila_ini - 1,
                        "col_0idx": col_0idx,
                        "bytes_png": bytes_png,
                        "ancho_px": ancho_f,
                        "alto_px": alto_f,
                    })
                else:
                    if prefijo_foto == "det" and key_id.startswith("esp_"):
                        ws.cell(row=fila_ini, column=12, value="-")

    fila_ot = 252 + desplazamiento
    fila_ot_max = 259 + desplazamiento
    for idx_ot, rechazo in enumerate(todos_los_rechazos):
        if fila_ot > fila_ot_max:
            break  
        defecto = rechazo["defecto"]
        prefijo = "SOLD_CBO" if defecto == "DE" else "SOLD_REP"
        codigo_sugerido = f"BK00{str(idx_ot + 1).zfill(5)}"
        codigo_backlog = st.session_state.get(f"bk_{rechazo['key_id']}", codigo_sugerido)
        texto_ot = f"{prefijo} {rechazo['descripcion']} ({rechazo['zona']}) - {codigo_backlog}"
        ws.cell(row=fila_ot, column=1, value=texto_ot)
        fila_ot += 1

    fila_nombre = 261 + desplazamiento
    fila_firma = 264 + desplazamiento
    fila_fecha = 266 + desplazamiento
    ws.cell(row=fila_nombre, column=2, value=nombre_realizado)
    ws.cell(row=fila_fecha, column=2, value=fecha_firma)

    if firma_archivo is not None:
        firma_archivo.seek(0)
        img_firma = Image.open(firma_archivo).convert("RGB")
        img_firma.thumbnail((220, 90), Image.LANCZOS)
        buf_firma = io.BytesIO()
        img_firma.save(buf_firma, format="PNG")
        fotos_pendientes.append({
            "fila_0idx": fila_firma - 1,
            "col_0idx": 1,  
            "bytes_png": buf_firma.getvalue(),
            "ancho_px": img_firma.width,
            "alto_px": img_firma.height,
        })

    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.scale = 50
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = False

    buf_temp = io.BytesIO()
    wb.save(buf_temp)
    buf_temp.seek(0)
    z_temp = zipfile.ZipFile(buf_temp)
    partes_nuevas = {}
    for nombre in ['xl/worksheets/sheet1.xml', 'xl/sharedStrings.xml', 'xl/styles.xml']:
        if nombre in z_temp.namelist():
            partes_nuevas[nombre] = z_temp.read(nombre)

    z_original = zipfile.ZipFile(ruta_plantilla)
    drawing1_xml = z_original.read('xl/drawings/drawing1.xml').decode('utf-8')
    drawing1_rels = z_original.read('xl/drawings/_rels/drawing1.xml.rels').decode('utf-8')

    for fila_insercion, cantidad in puntos_insercion:
        drawing1_xml = _desplazar_filas_drawing_xml(drawing1_xml, fila_insercion - 1, cantidad)

    import re
    rids_existentes = re.findall(r'Id="rId(\d+)"', drawing1_rels)
    siguiente_rid = max((int(r) for r in rids_existentes), default=0) + 1
    media_nuevos = {}
    anchors_nuevos_xml = ""
    id_shape = 9000

    for foto in fotos_pendientes:
        nombre_media = f"appFoto{siguiente_rid}.png"
        media_nuevos[f"xl/media/{nombre_media}"] = foto["bytes_png"]
        rid_actual = f"rId{siguiente_rid}"
        drawing1_rels = drawing1_rels.replace(
            "</Relationships>",
            f'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="../media/{nombre_media}" Id="{rid_actual}"/></Relationships>'
        )
        anchors_nuevos_xml += _construir_anchor_imagen_xml(
            id_shape, foto["col_0idx"], foto["fila_0idx"], foto["ancho_px"], foto["alto_px"], rid_actual
        )
        siguiente_rid += 1
        id_shape += 1

    drawing1_xml = drawing1_xml.replace("</xdr:wsDr>", anchors_nuevos_xml + "</xdr:wsDr>")
    partes_nuevas['xl/drawings/drawing1.xml'] = drawing1_xml.encode('utf-8')
    partes_nuevas['xl/drawings/_rels/drawing1.xml.rels'] = drawing1_rels.encode('utf-8')

    buf_final = io.BytesIO()
    with zipfile.ZipFile(buf_final, 'w', zipfile.ZIP_DEFLATED) as z_final:
        for item in z_original.infolist():
            datos = partes_nuevas.get(item.filename, z_original.read(item.filename))
            z_final.writestr(item, datos)
        for nombre_media, datos_media in media_nuevos.items():
            z_final.writestr(nombre_media, datos_media)
    buf_final.seek(0)
    return buf_final.getvalue(), []

def convertir_excel_a_pdf(bytes_excel):
    import subprocess
    import tempfile

    with tempfile.TemporaryDirectory() as carpeta_temp:
        ruta_xlsx = os.path.join(carpeta_temp, "reporte.xlsx")
        with open(ruta_xlsx, "wb") as f:
            f.write(bytes_excel)
        try:
            subprocess.run(
                ["soffice", "--headless", "--convert-to", "pdf", "--outdir", carpeta_temp, ruta_xlsx],
                check=True, timeout=120, capture_output=True
            )
        except Exception:
            return None

        ruta_pdf = os.path.join(carpeta_temp, "reporte.pdf")
        if os.path.exists(ruta_pdf):
            with open(ruta_pdf, "rb") as f:
                return f.read()
        return None

def redimensionar_conservando_calidad(img, max_lado=1400):
    ancho, alto = img.size
    escala = min(max_lado / max(ancho, alto), 1.0)
    if escala >= 1.0:
        return img
    nuevo_ancho = max(1, int(ancho * escala))
    nuevo_alto = max(1, int(alto * escala))
    return img.resize((nuevo_ancho, nuevo_alto), Image.LANCZOS)

def mostrar_imagen_responsive(ruta_o_objeto, caption=None):
    try:
        st.image(ruta_o_objeto, caption=caption, width="stretch")
    except TypeError:
        try:
            st.image(ruta_o_objeto, caption=caption, use_container_width=True)
        except TypeError:
            try:
                st.image(ruta_o_objeto, caption=caption, use_column_width=True)
            except TypeError:
                st.image(ruta_o_objeto, caption=caption)

# --- BASE DE DATOS LOCAL PARA RECORDAR TOLVA ---
DB_FILE = os.path.join("base_datos", "tolvas_db.json")

def cargar_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "r") as f:
            try:
                return json.load(f)
            except:
                return {}
    return {}

def guardar_db(data):
    os.makedirs("base_datos", exist_ok=True)
    with open(DB_FILE, "w") as f:
        json.dump(data, f, indent=4)

db_tolvas = cargar_db()

# --- SECCIÓN 1: ENCABEZADO ---
st.header("1. Datos Generales del Informe")

col1, col2, col3 = st.columns(3)

with col1:
    opc_cliente = st.selectbox("Cliente:", ["ANGLOAMERICAN QUELLAVECO S.A.", "[Entrada Manual]"], key="header_cliente_select")
    cliente = st.text_input("Nombre del Cliente:", value="ANGLOAMERICAN QUELLAVECO S.A.", key="header_cliente_input") if opc_cliente == "[Entrada Manual]" else opc_cliente

    opc_lugar = st.selectbox("Lugar:", ["TRUCK SHOP", "[Entrada Manual]"], key="header_lugar_select")
    lugar = st.text_input("Lugar de Inspección:", value="TRUCK SHOP", key="header_lugar_input") if opc_lugar == "[Entrada Manual]" else opc_lugar

    fecha_insp = st.date_input("Fecha de Inspección:", value=date.today(), key="header_fecha")

with col2:
    lista_equipos = [f"HT{str(i).zfill(3)}" for i in range(1, 35)]
    cod_equipo = st.selectbox("Código de Equipo:", lista_equipos, index=1, key="header_equipo")

    tolva_recordada = db_tolvas.get(cod_equipo, "T-CA2" if cod_equipo == "HT002" else "")
    cod_tolva = st.text_input("Código de Tolva:", value=tolva_recordada, key="header_tolva")

    if cod_tolva and cod_tolva != db_tolvas.get(cod_equipo):
        db_tolvas[cod_equipo] = cod_tolva
        guardar_db(db_tolvas)

    horometro = st.number_input("Horómetro (hrs):", value=34876.3, step=0.1, format="%.1f", key="header_horometro")

with col3:
    sufijo_informe = st.text_input("Sufijo del Informe (XXX):", value="023", key="header_sufijo")
    cod_informe = f"ZA-IF-{cod_equipo}-{sufijo_informe}"
    st.info(f"Código Generado: **{cod_informe}**")

    revision = st.text_input("Revisión:", value="00", key="header_revision")
    pm = st.selectbox("Mantenimiento (PM):", ["500H", "1000H", "1500H", "2000H"], key="header_pm")

st.markdown("---")

# --- SECCIÓN MEDICIÓN DE ESPESORES (Solo en 1000H y 2000H) ---
matriz_espesores_final = None

if pm in ["1000H", "2000H"]:
    st.header("📏 Mapeo / Medición de Espesores (Matriz 8x7)")
    no_medicion = st.checkbox("⚠️ Marcar como 'NO SE REALIZÓ MEDICIÓN DE ESPESORES'", key="check_no_medicion")

    if no_medicion:
        st.warning("Se ha seleccionado omitir la medición. El reporte se llenará con '-' y agregará la nota explicativa.")
        matriz_espesores_final = pd.DataFrame([["-"]*7 for _ in range(8)], 
                                    index=[f"Punto {i+1}" for i in range(8)],
                                    columns=[f"Eje {j+1}" for j in range(7)])
        st.dataframe(matriz_espesores_final, use_container_width=True)
    else:
        st.caption("Ingrese manualmente las lecturas de ultrasonido (mm) en la matriz:")
        df_init_espesores = pd.DataFrame([[20.00]*7 for _ in range(8)], 
                               index=[f"Punto {i+1}" for i in range(8)],
                               columns=[f"Eje {j+1}" for j in range(7)])
        
        column_config_espesores = {
            f"Eje {i+1}": st.column_config.NumberColumn(width=65, format="%.2f") for i in range(7)
        }

        matriz_espesores_final = st.data_editor(
            df_init_espesores, 
            use_container_width=False,
            column_config=column_config_espesores,
            key="editor_espesores"
        )

    st.markdown("---")

# --- ESQUEMA VISUAL GENERAL ---
if os.path.exists("esquema_tolva.png"):
    st.subheader("🗺️ Esquema Guía General de Zonas")
    mostrar_imagen_responsive("esquema_tolva.png", caption="Plano de Ubicación General de Componentes - Tolva CAT 794 AC")
elif os.path.exists("esquema_tolva.jpg"):
    st.subheader("🗺️ Esquema Guía General de Zonas")
    mostrar_imagen_responsive("esquema_tolva.jpg", caption="Plano de Ubicación General de Componentes - Tolva CAT 794 AC")

# --- CONFIGURACIÓN DE ESTRUCTURA DE ZONAS ---
ESTRUCTURA_ZONAS = [
    {
        "titulo": "ZONA 01: CONJUNTO DE BLINDAJE DE TOLVA",
        "esquema": ["CONJUNTO DE BLINDAJE DE TOLVA.png"],
        "items": [
            ("1.1", "REFUERZO DE PISO #1", "VT/UT"),
            ("1.2", "REFUERZO DE PISO #2", "VT/UT"),
            ("1.3", "REFUERZO DE PISO #3", "VT/UT"),
            ("1.4", "ROCKBOX", "VT"),
            ("1.5", "REFUERZO FRONTAL #1", "VT/UT"),
            ("1.6", "REFUERZO FRONTAL #2", "VT/UT"),
            ("1.7", "REFUERZO FRONTAL #3", "VT/UT"),
            ("1.8", "REFUERZO LATERAL RH", "VT/UT"),
            ("1.9", "REFUERZO LATERAL LH", "VT"),
            ("1.10", "CORTAFLUJOS", "VT/UT")
        ]
    },
    {
        "titulo": "ZONA 02: CONJUNTO LATERAL (RH / LH)",
        "esquema": ["CONJUNTO LATERAL.png"],
        "items": [
            ("2.1", "PLANCHA LATERAL RH", "VT"),
            ("2.2", "VIGA CAJON LATERAL RH", "VT"),
            ("2.3", "PLANCHA LATERAL LH", "VT"),
            ("2.4", "VIGA CAJON LATERAL LH", "VT")
        ]
    },
    {
        "titulo": "ZONA 03: CANOPY",
        "esquema": ["CANOPY.png"],
        "items": [
            ("3.1", "PLANCHA RH", "VT"),
            ("3.2", "PLANCHA LH", "VT"),
            ("3.3", "DEFLECTOR RH", "VT"),
            ("3.4", "DEFLECTOR LH", "VT"),
            ("3.5", "PLANCHA FRONTAL CANOPY", "VT"),
            ("3.6", "CARTELAS DE PLANCHA FRONTAL", "VT"),
            ("3.7", "VIGA LATERAL RH", "VT"),
            ("3.8", "VIGA LATERAL LH", "VT"),
            ("3.9", "REFUERZO RH DE CANOPY", "VT"),
            ("3.10", "REFUERZO LH DE CANOPY", "VT")
        ]
    },
    {
        "titulo": "ZONA 04 Y 05: PLANCHAS DE PISO / PLANCHAS FRONTALES",
        "esquema": ["PLANCHAS FRONTALES.png", "PLANCHAS DE PISO.png"],
        "items": [
            ("4.1", "PLANCHA FRONTAL SUPERIOR", "VT"),
            ("4.2", "PLANCHA FRONTAL RH", "VT"),
            ("4.3", "PLANCHA FRONTAL LH", "VT"),
            ("5.1", "PLANCHA DE PISO RH", "VT/UT"),
            ("5.2", "PLANCHA DE PISO LH", "VT/UT"),
            ("5.3", "PLANCHA COLA DE PISO", "VT/UT")
        ]
    },
    {
        "titulo": "ZONA 06 Y 07: LONGUERINA DELANTERA, POSTERIOR Y VIGAS / GUIADORES",
        "esquema": ["LONGUERINA DELANTERA, POSTERIOR Y VIGAS.png", "GUIADORES.png"],
        "items": [
            ("6.1", "LONGUERINA DELANTERA RH", "VT"),
            ("6.2", "LONGUERINA DELANTERA LH", "VT"),
            ("6.3", "LONGUERINA POSTERIOR RH", "VT"),
            ("6.4", "LONGUERINA POSTERIOR LH", "VT"),
            ("6.5", "VIGA DE PISO #1 RH", "VT"),
            ("6.6", "VIGA CAJON DE PISO #1 LH", "VT"),
            ("6.7", "VIGA CAJON DE PISO #2 RH", "VT"),
            ("6.8", "VIGA CAJON DE PISO #2 LH", "VT"),
            ("6.9", "VIGA CAJON DE COLA", "VT"),
            ("7.1", "GUIADOR RH", "VT"),
            ("7.2", "GUIADOR LH", "VT")
        ]
    },
    {
        "titulo": "ZONA 08: CAJAS PIVOTE",
        "esquema": ["CAJAS PIVOTE.png"],
        "items": [
            ("8.1", "CAJA PIVOTE RH", "VT"),
            ("8.2", "CAJA PIVOTE LH", "VT"),
            ("8.3", "BUSHING DE CAJA PIVOTE RH", "VT"),
            ("8.4", "BUSHING DE CAJA PIVOTE LH", "VT"),
            ("8.5", "SEPARADOR DE CAJA PIVOTE", "VT")
        ]
    }
]

def mostrar_esquema_zona(nombres_archivo, titulo_zona):
    if isinstance(nombres_archivo, str):
        nombres_archivo = [nombres_archivo]

    rutas_existentes = [
        os.path.join("imagenes_esquemas", n) for n in nombres_archivo
        if os.path.exists(os.path.join("imagenes_esquemas", n))
    ]

    if rutas_existentes:
        st.markdown("**🗺️ Esquema de referencia de la zona:**")
        if len(rutas_existentes) == 1:
            mostrar_imagen_responsive(rutas_existentes[0])
        else:
            cols = st.columns(len(rutas_existentes))
            for c, ruta in zip(cols, rutas_existentes):
                with c:
                    mostrar_imagen_responsive(ruta)
        st.markdown("")

# --- GESTOR FOTOGRÁFICO ---
def gestor_fotografico(label_foto, key_foto):
    st.markdown(f"**{label_foto}**")

    llave_img = f"img_{key_foto}"
    llave_anotada = f"{llave_img}_anotada"

    if f"retomar_{key_foto}" in st.session_state and st.session_state[f"retomar_{key_foto}"]:
        for k in [llave_img, llave_anotada]:
            if k in st.session_state:
                del st.session_state[k]
        del st.session_state[f"retomar_{key_foto}"]

    if llave_img not in st.session_state:
        metodo = st.radio(
            f"Modo de Carga ({key_foto}):", 
            ["Galería / Archivo", "Cámara Directa"], 
            key=f"rad_{key_foto}", 
            horizontal=True
        )
        
        img_upload = None
        if metodo == "Cámara Directa":
            foto_b64 = camara_nativa(key=f"camnat_{key_foto}")
            if foto_b64:
                datos_binarios = base64.b64decode(foto_b64.split(",", 1)[1])
                img_upload = Image.open(io.BytesIO(datos_binarios))
        else:
            up_data = st.file_uploader(f"Subir {label_foto}", type=["jpg", "png", "jpeg"], key=f"up_{key_foto}")
            if up_data:
                img_upload = Image.open(up_data)

        if img_upload is not None:
            img_fija = ImageOps.exif_transpose(img_upload.convert("RGB"))
            img_resized = redimensionar_conservando_calidad(img_fija, max_lado=1400)
            st.session_state[llave_img] = img_resized
            st.rerun()

    if llave_img in st.session_state:
        img_actual = st.session_state[llave_img]

        if st.button(f"🗑️ Retomar Foto", key=f"btn_retomar_{key_foto}"):
            st.session_state[f"retomar_{key_foto}"] = True
            st.rerun()

        buf = io.BytesIO()
        img_actual.save(buf, format="JPEG", quality=90)
        img_b64 = base64.b64encode(buf.getvalue()).decode()

        resultado = anotador_fotos(img_b64, key=f"anot_{key_foto}")
        if resultado:
            st.session_state[llave_anotada] = resultado

        if llave_anotada in st.session_state:
            st.caption("✅ Anotación guardada para esta foto.")

# --- PROCESAMIENTO Y DESPLIEGUE DE ZONAS ---
todos_los_rechazos = [] 

for idx_z, bloque_zona in enumerate(ESTRUCTURA_ZONAS):
    st.header(bloque_zona["titulo"])
    mostrar_esquema_zona(bloque_zona["esquema"], bloque_zona["titulo"])

    h1, h2, h3, h4, h5, h6, h7, h8, h9 = st.columns([0.6, 2.2, 1.2, 0.8, 1.0, 1.0, 1.0, 1.2, 1.2])
    with h1: st.markdown("**ZONA**")
    with h2: st.markdown("**DESCRIPCIÓN**")
    with h3: st.markdown("**FECHA**")
    with h4: st.markdown("**DEFECTO**")
    with h5: st.markdown("**LONG. (mm)**")
    with h6: st.markdown("**EST. POST.**")
    with h7: st.markdown("**TÉCNICA**")
    with h8: st.markdown("**CONDICIÓN**")
    with h9: st.markdown("**COMENTARIOS**")

    rechazos_de_esta_zona = []

    for idx_i, item in enumerate(bloque_zona["items"]):
        cod_z, desc_z, tec_def = item
        key_id = f"z{idx_z}_{idx_i}"

        c1, c2, c3, c4, c5, c6, c7, c8, c9 = st.columns([0.6, 2.2, 1.2, 0.8, 1.0, 1.0, 1.0, 1.2, 1.2])

        with c1:
            st.caption("ZONA")
            st.write(f"**{cod_z}**")
        with c2:
            st.caption("DESCRIPCIÓN")
            st.write(desc_z)
        with c3:
            st.caption("FECHA")
            st.write(fecha_insp.strftime("%d/%m/%Y"))

        with c4:
            defecto = st.selectbox("DEFECTO", ["LF", "D", "DE", "DP", "F", "FA"], key=f"def_{key_id}")

        es_lf = (defecto == "LF")

        with c5:
            if es_lf:
                longitud = "-"
                st.text_input("LONG. (mm)", value="-", disabled=True, key=f"long_{key_id}")
            else:
                opc_long = st.selectbox("LONG. (mm)", ["Manual", "VARIOS"], key=f"opclong_{key_id}")
                if opc_long == "VARIOS":
                    longitud = "VARIOS"
                else:
                    longitud = st.text_input("Valor longitud", value="100", key=f"longval_{key_id}")

        with c6:
            if es_lf:
                est_post = "-"
                st.selectbox("EST. POST.", ["-"], disabled=True, key=f"est_{key_id}")
            else:
                est_post = st.selectbox("EST. POST.", ["NR", "R"], key=f"est_{key_id}")

        with c7:
            tecnica = st.selectbox("TÉCNICA", ["VT", "VT/PT", "VT/UT"], index=2 if tec_def=="VT/UT" else 0, key=f"tec_{key_id}")

        with c8:
            st.caption("CONDICIÓN")
            if es_lf:
                condicion = "ACEPTABLE"
                st.markdown("<span style='color:#48BB78; font-weight:bold;'>ACEPTABLE</span>", unsafe_allow_html=True)
            else:
                condicion = "RECHAZADO"
                st.markdown("<span style='color:#F56565; font-weight:bold;'>RECHAZADO</span>", unsafe_allow_html=True)

        with c9:
            if es_lf:
                comentario = "-"
                st.text_input("COMENTARIOS", value="-", disabled=True, key=f"com_{key_id}")
            else:
                comentario = st.selectbox("COMENTARIOS", ["CREAR OT", "OT CREADA"], key=f"com_{key_id}")

        if not es_lf:
            datos_defectuosos = {
                "zona": cod_z,
                "descripcion": desc_z,
                "defecto": defecto,
                "longitud": longitud,
                "est_post": est_post,
                "tecnica": tecnica,
                "comentario": comentario,
                "key_id": key_id
            }
            rechazos_de_esta_zona.append(datos_defectuosos)
            todos_los_rechazos.append(datos_defectuosos)

    if len(rechazos_de_esta_zona) > 0:
        st.subheader(f"📸 Registro Fotográfico de Defectos - {bloque_zona['titulo']}")
        for rechazo in rechazos_de_esta_zona:
            expander_label = f"📷 Fotos de Hallazgo: {rechazo['descripcion']} ({rechazo['zona']}) - Defecto: [{rechazo['defecto']}]"
            with st.expander(expander_label, expanded=True):
                f_col1, f_col2 = st.columns(2)
                with f_col1:
                    gestor_fotografico("Foto Panorámica", f"pano_{rechazo['key_id']}")
                with f_col2:
                    gestor_fotografico("Foto de Detalle", f"det_{rechazo['key_id']}")

    if "ZONA 02" in bloque_zona["titulo"]:
        with st.expander("📷 Letreros Obligatorios Especiales: Laterales RH / LH", expanded=False):
            f_col1, f_col2 = st.columns(2)
            with f_col1:
                gestor_fotografico("Letrero Lateral RH (Panorámico)", "pano_esp_z2_rh")
            with f_col2:
                gestor_fotografico("Letrero Lateral LH (Panorámico)", "pano_esp_z2_lh")
                
    elif "ZONA 08" in bloque_zona["titulo"]:
        with st.expander("📷 Letrero Obligatorio Especial: Posterior", expanded=False):
            f_col1, f_col2 = st.columns(2)
            with f_col1:
                gestor_fotografico("Letrero Posterior (Panorámico)", "pano_esp_z8_post")
            with f_col2:
                st.info("Requerido para la zona posterior (Z08). El campo 'Detalle' en el Excel se llenará automáticamente con '-'.")

    st.caption("**Leyenda Defectos:** D: Desprendimiento | DE: Desgaste | DP: Desprendimiento Parcial | F: Fisurado | LF: Libre de fisura (Aceptable)")
    st.caption("**Leyenda Técnica:** NR: No Reparado | R: Reparado | VT: Inspección Visual | LP: Líquidos Penetrantes | UT: Ultrasonido")
    st.markdown("---")

# --- SECCIÓN 4: ZONAS A REPARAR ---
st.header("4. Resumen de Órdenes de Trabajo (Zonas a Reparar)")

if len(todos_los_rechazos) == 0:
    st.success("✔ No se registraron defectos. No hay reparaciones pendientes requeridas.")
else:
    st.caption("Resumen sugerido de Órdenes de Trabajo según los defectos reportados:")
    for idx_ot, rechazo in enumerate(todos_los_rechazos):
        defecto = rechazo["defecto"]
        prefijo = "SOLD_CBO" if defecto == "DE" else "SOLD_REP"
        codigo_sugerido = f"BK00{str(idx_ot + 1).zfill(5)}"
        
        col_ot1, col_ot2 = st.columns([3, 2])
        with col_ot1:
            st.code(f"{prefijo} {rechazo['descripcion']} ({rechazo['zona']}) - {rechazo['defecto']}")
        with col_ot2:
            st.text_input(
                f"Código Backlog SAP ({rechazo['zona']}):", 
                value=codigo_sugerido, 
                key=f"bk_{rechazo['key_id']}"
            )

st.markdown("---")
st.header("5. Firma del Responsable de la Inspección")

col_f1, col_f2 = st.columns(2)
with col_f1:
    nombre_realizado = st.text_input("Nombre de quien realiza la inspección:", key="firma_nombre")
    fecha_firma = st.date_input("Fecha de firma:", value=fecha_insp, key="firma_fecha")
with col_f2:
    firma_archivo = st.file_uploader("Subir imagen de firma (PNG/JPG):", type=["png", "jpg", "jpeg"], key="firma_upload")
    if firma_archivo:
        st.image(firma_archivo, caption="Vista previa de la firma", width=250)

st.markdown("---")
st.header("6. Generar Reporte Final")

RUTA_PLANTILLA = "plantilla_tolva.xlsx"
nombre_archivo_base = f"{cod_informe}_{fecha_insp.strftime('%Y%m%d')}"

if not os.path.exists(RUTA_PLANTILLA):
    st.error(
        f"⚠️ No se encontró el archivo '{RUTA_PLANTILLA}' en el proyecto. "
        "Sube la plantilla original de Excel a la misma carpeta que app.py en GitHub "
        "(con ese nombre exacto) para poder generar el reporte."
    )
else:
    if st.button("📥 Generar Reporte", type="primary"):
        with st.spinner("Generando el archivo Excel..."):
            excel_bytes, zonas_sin_espacio = generar_reporte_excel(
                ruta_plantilla=RUTA_PLANTILLA,
                cliente=cliente, lugar=lugar, fecha_insp=fecha_insp,
                cod_equipo=cod_equipo, cod_tolva=cod_tolva, horometro=horometro,
                cod_informe=cod_informe, revision=revision, pm=pm,
                estructura_zonas=ESTRUCTURA_ZONAS,
                nombre_realizado=nombre_realizado, fecha_firma=fecha_firma,
                firma_archivo=firma_archivo,
                todos_los_rechazos=todos_los_rechazos,
                matriz_espesores=matriz_espesores_final
            )
        st.session_state["_ultimo_excel_generado"] = excel_bytes
        st.success("✅ Reporte Excel generado correctamente.")

    if "_ultimo_excel_generado" in st.session_state:
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="💾 Descargar Excel",
                data=st.session_state["_ultimo_excel_generado"],
                file_name=f"{nombre_archivo_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col_dl2:
            if st.button("📄 Convertir y Descargar PDF"):
                with st.spinner("Convirtiendo a PDF... esto puede tardar un poco"):
                    pdf_bytes = convertir_excel_a_pdf(st.session_state["_ultimo_excel_generado"])
                if pdf_bytes:
                    st.download_button(
                        label="💾 Descargar PDF",
                        data=pdf_bytes,
                        file_name=f"{nombre_archivo_base}.pdf",
                        mime="application/pdf"
                    )
                else:
                    st.error(
                        "⚠️ No se pudo convertir a PDF. Esto pasa si el servidor no tiene "
                        "LibreOffice instalado (revisa que subiste el archivo 'packages.txt'). "
                        "Mientras tanto, puedes abrir el Excel descargado y usar "
                        "'Archivo > Exportar > Crear PDF/XPS' desde Excel o Google Sheets."
                    )

        st.caption(
            "✅ Los logos y el formato original de la empresa se conservan intactos. "
            "Si una zona tiene más de 2 hallazgos, se crean bloques de foto adicionales "
            "automáticamente (mismo estilo/bordes que la plantilla). "
            "Las zonas 2 y 8 tienen sus espacios de letreros fijos, se suban o no fotos."
        )
